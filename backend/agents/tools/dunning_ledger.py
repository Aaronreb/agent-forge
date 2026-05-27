import os
from datetime import datetime
from langchain_core.tools import tool
from openpyxl import load_workbook, Workbook

_LEDGER_PATH = "/home/konsultera/Aaron/Projects/agentplatform/dunning_ledger.xlsx"
_HEADERS = [
    "Subscription ID", "Transaction ID", "Timestamp",
    "Decline Code", "Decline Type", "LTV Tier", "Tenure Months",
    "Retry Count", "Action Taken", "Notification Sent", "Outcome",
]


@tool
def dunning_ledger_tool(
    subscription_id: str,
    transaction_id: str,
    decline_code: str,
    decline_type: str,
    ltv_tier: str,
    tenure_months: str,
    retry_count: str,
    action_taken: str,
    notification_sent: str,
    outcome: str,
) -> str:
    """Log a dunning event to the Excel dunning ledger.
    Call this as the final step in both recovery and cancellation paths.
    action_taken: e.g. 'retry_scheduled_3d', 'plan_paused_30d', 'subscription_cancelled'
    outcome: e.g. 'retry_pending', 'paused_30d', 'cancelled'
    notification_sent: the template name used (e.g. 'friendly_reminder')."""
    try:
        if os.path.exists(_LEDGER_PATH):
            wb = load_workbook(_LEDGER_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Dunning Ledger"
            ws.append(_HEADERS)

        ws.append([
            subscription_id,
            transaction_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            decline_code,
            decline_type,
            ltv_tier,
            tenure_months,
            retry_count,
            action_taken,
            notification_sent,
            outcome,
        ])
        wb.save(_LEDGER_PATH)
        return f"Dunning event logged at row {ws.max_row} | sub={subscription_id} | action={action_taken}"
    except Exception as e:
        return f"Dunning ledger write failed: {e}"
