import os
from datetime import datetime
from langchain_core.tools import tool
from openpyxl import load_workbook, Workbook

_EXCEL_PATH = "/home/konsultera/Aaron/Projects/agentplatform/support_tickets.xlsx"
_HEADERS = ["Ticket ID", "Timestamp", "Category", "Priority", "Sentiment", "Summary", "KB Articles Used"]


@tool
def ticket_logger(
    ticket_id: str,
    category: str,
    priority: str,
    sentiment: str,
    summary: str,
    kb_articles_used: str = "",
) -> str:
    """Log a classified support ticket to the central Excel tracker.
    Call this after classifying the ticket and finding KB articles.
    ticket_id: a short identifier (e.g. 'TKT-001' or first 8 chars of a UUID).
    kb_articles_used: comma-separated article IDs referenced (e.g. 'KB001, KB003').
    Returns confirmation with the row number written."""
    try:
        if os.path.exists(_EXCEL_PATH):
            wb = load_workbook(_EXCEL_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Support Tickets"
            ws.append(_HEADERS)

        ws.append([
            ticket_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            category,
            priority,
            sentiment,
            summary,
            kb_articles_used,
        ])
        wb.save(_EXCEL_PATH)
        return f"Ticket {ticket_id} logged at row {ws.max_row} in {_EXCEL_PATH}"
    except Exception as e:
        return f"Excel logging failed: {e}"
